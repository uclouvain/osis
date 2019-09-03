##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 UniversitÃ© catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import itertools
from collections import namedtuple, defaultdict

from django.db.models import QuerySet, Prefetch
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Style, Border, Side, Color, PatternFill, Font
from openpyxl.styles.borders import BORDER_THICK
from openpyxl.styles.colors import RED
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.writer.excel import save_virtual_workbook

from backoffice.settings.base import LEARNING_UNIT_PORTAL_URL
from base.models.education_group_year import EducationGroupYear
from base.models.enums.prerequisite_operator import OR
from base.models.group_element_year import fetch_row_sql, GroupElementYear
from base.models.learning_unit_year import LearningUnitYear
from base.models.prerequisite import Prerequisite
from base.models.prerequisite_item import PrerequisiteItem
from osis_common.document.xls_build import _build_worksheet, CONTENT_KEY, HEADER_TITLES_KEY, WORKSHEET_TITLE_KEY, \
    STYLED_CELLS, STYLE_NO_GRAY

STYLE_BORDER_BOTTOM = Style(
    border=Border(
        bottom=Side(
            border_style=BORDER_THICK, color=Color('FF000000')
        )
    )
)
STYLE_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('D1D1D1')))
STYLE_LIGHT_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('E1E1E1')))
STYLE_LIGHTER_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('F1F1F1')))

STYLE_FONT_RED = Style(font=Font(color=RED))
FONT_HYPERLINK = Font(underline='single', color='0563C1')

HeaderLine = namedtuple('HeaderLine', ['egy_acronym', 'egy_title'])
OfficialTextLine = namedtuple('OfficialTextLine', ['text'])
LearningUnitYearLine = namedtuple('LearningUnitYearLine', ['luy_acronym', 'luy_title'])
PrerequisiteItemLine = namedtuple(
    'PrerequisiteItemLine',
    ['text', 'operator', 'luy_acronym', 'luy_title', 'credits', 'block', 'mandatory']
)


class EducationGroupYearLearningUnitsPrerequisitesToExcel:
    def __init__(self, egy: EducationGroupYear):
        self.egy = egy

    def get_queryset(self):
        group_element_years_of_education_group_year = [element["id"] for element in fetch_row_sql([self.egy.id])]
        return Prerequisite.objects.filter(
            education_group_year=self.egy
        ).prefetch_related(
            Prefetch(
                "prerequisiteitem_set",
                queryset=PrerequisiteItem.objects.order_by(
                    'group_number',
                    'position'
                ).select_related(
                    "learning_unit"
                ).prefetch_related(
                    Prefetch(
                        "learning_unit__learningunityear_set",
                        queryset=LearningUnitYear.objects.filter(
                            academic_year=self.egy.academic_year
                        ).prefetch_related(
                            Prefetch(
                                "child_leaf",
                                queryset=GroupElementYear.objects.filter(
                                    child_leaf__isnull=False,
                                    id__in=group_element_years_of_education_group_year
                                ),
                                to_attr="links"
                            )
                        ),
                        to_attr="luys"
                    )
                ),
                to_attr="items"
            )
        ).select_related(
            "learning_unit_year"
        ).order_by(
            "learning_unit_year__acronym"
        )

    def _to_workbook(self):
        return generate_prerequisites_workbook(self.egy, self.get_queryset())

    def to_excel(self):
        return save_virtual_workbook(self._to_workbook())


def generate_prerequisites_workbook(egy: EducationGroupYear, prerequisites_qs: QuerySet):
    worksheet_title = _("prerequisites-%(year)s-%(acronym)s") % {"year": egy.academic_year.year, "acronym": egy.acronym}
    workbook = Workbook(encoding='utf-8')

    excel_lines = _build_excel_lines(egy, prerequisites_qs)

    header, *content = [tuple(line) for line in excel_lines]
    style = _get_style_to_apply(excel_lines)

    worksheet_data = {
        WORKSHEET_TITLE_KEY: worksheet_title,
        HEADER_TITLES_KEY: header,
        CONTENT_KEY: content,
        STYLED_CELLS: style
    }
    _build_worksheet(worksheet_data, workbook, 0)

    _merge_cells(excel_lines, workbook)
    _add_hyperlink(excel_lines, workbook, str(egy.academic_year.year))
    return workbook


def _build_excel_lines(egy: EducationGroupYear, prerequisite_qs: QuerySet):
    content = []
    content.append(
        HeaderLine(egy_acronym=egy.acronym, egy_title=egy.title)
    )
    content.append(
        OfficialTextLine(text=_("Official"))
    )

    for prerequisite in prerequisite_qs:
        luy = prerequisite.learning_unit_year
        content.append(
            LearningUnitYearLine(luy_acronym=luy.acronym, luy_title=luy.complete_title_i18n)
        )
        groups_generator = itertools.groupby(prerequisite.items, key=lambda item: item.group_number)
        for key, group_gen in groups_generator:

            group = list(group_gen)
            for item in group:
                prerequisite_line = _build_prerequisite_line(prerequisite, item, group)
                content.append(prerequisite_line)

    return content


def _build_prerequisite_line(prerequisite: Prerequisite, prerequisite_item: PrerequisiteItem, group: list):
    luy_item = prerequisite_item.learning_unit.luys[0]

    text = (_("has as prerequisite") + " :") \
        if prerequisite_item.group_number == 1 and prerequisite_item.position == 1 else None
    operator = _get_operator(prerequisite, prerequisite_item)
    luy_acronym = _get_item_acronym(prerequisite_item, group)
    credits = _get_item_credits(prerequisite_item)
    block = _get_item_blocks(prerequisite_item)
    mandatory = luy_item.links[0].is_mandatory if luy_item.links else None
    return PrerequisiteItemLine(
        text=text,
        operator=operator,
        luy_acronym=luy_acronym,
        luy_title=prerequisite_item.learning_unit.luys[0].complete_title_i18n,
        credits=credits,
        block=block,
        mandatory=_("Yes") if mandatory else _("No")
    )


def _get_operator(prerequisite: Prerequisite, prerequisite_item: PrerequisiteItem):
    if prerequisite_item.group_number == 1 and prerequisite_item.position == 1:
        return None
    elif prerequisite_item.position == 1:
        return _(prerequisite.main_operator)
    return _(prerequisite.secondary_operator)


def _get_item_acronym(prerequisite_item: PrerequisiteItem, group: list):
    acronym_format = "{acronym}"
    if prerequisite_item.position == 1 and len(group) > 1:
        acronym_format = "({acronym}"
    elif prerequisite_item.position == len(group) and len(group) > 1:
        acronym_format = "{acronym})"
    return acronym_format.format(acronym=prerequisite_item.learning_unit.luys[0].acronym)


def _get_item_credits(prerequisite_item: PrerequisiteItem):
    luy_item = prerequisite_item.learning_unit.luys[0]
    return " ; ".join(
        set(["{} / {:f}".format(grp.relative_credits, luy_item.credits.normalize()) for grp in luy_item.links])
    )


def _get_item_blocks(prerequisite_item: PrerequisiteItem):
    luy_item = prerequisite_item.learning_unit.luys[0]
    return " ; ".join(
        [str(grp.block) for grp in luy_item.links if grp.block]
    )


def _get_style_to_apply(excel_lines: list):
    style_to_apply_dict = defaultdict(list)
    last_luy_line_index = None
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, HeaderLine):
            style_to_apply_dict[STYLE_NO_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("B{index}".format(index=index))

        elif isinstance(row, OfficialTextLine):
            style_to_apply_dict[STYLE_BORDER_BOTTOM].append("A{index}".format(index=index))

        elif isinstance(row, LearningUnitYearLine):
            style_to_apply_dict[STYLE_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[STYLE_LIGHT_GRAY].append("B{index}".format(index=index))
            last_luy_line_index = index

        elif isinstance(row, PrerequisiteItemLine):
            if row.operator == _(OR):
                style_to_apply_dict[STYLE_FONT_RED].append("B{index}".format(index=index))

            if (last_luy_line_index - index) % 2 == 1:
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("C{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("D{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("E{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("F{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("G{index}".format(index=index))

    return style_to_apply_dict


def _merge_cells(excel_lines, workbook: Workbook):
    worksheet = workbook.worksheets[0]
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, LearningUnitYearLine):
            worksheet.merge_cells(start_row=index, end_row=index, start_column=2, end_column=7)


def _add_hyperlink(excel_lines, workbook: Workbook, year):
    worksheet = workbook.worksheets[0]
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, LearningUnitYearLine):
            cell = worksheet.cell(row=index, column=1)
            cell.hyperlink = LEARNING_UNIT_PORTAL_URL.format(year=year, acronym=row.luy_acronym)
            cell.font = FONT_HYPERLINK

        if isinstance(row, PrerequisiteItemLine):
            cell = worksheet.cell(row=index, column=3)
            cell.hyperlink = LEARNING_UNIT_PORTAL_URL.format(year=year, acronym=row.luy_acronym.strip("()"))
            cell.font = FONT_HYPERLINK
