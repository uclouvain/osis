##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import itertools
from abc import ABC, abstractmethod
from collections import namedtuple

from django.db.models import QuerySet, Prefetch
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Style, Border, Side, Color, PatternFill, Font
from openpyxl.styles.borders import BORDER_THICK
from openpyxl.styles.colors import RED
from openpyxl.writer.excel import save_virtual_workbook

from base.models.education_group_year import EducationGroupYear
from base.models.enums.prerequisite_operator import AND, OR
from base.models.learning_unit_year import LearningUnitYear
from base.models.prerequisite import Prerequisite
from base.models.prerequisite_item import PrerequisiteItem
from osis_common.document.xls_build import _build_worksheet, CONTENT_KEY, HEADER_TITLES_KEY, WORKSHEET_TITLE_KEY, \
    STYLED_CELLS, STYLE_NO_GRAY

XLS_FILENAME = "Formation prerequisites"
STYLE_BORDER_BOTTOM = Style(
    border=Border(
        bottom=Side(
            border_style=BORDER_THICK, color=Color('FF000000')
        )
    )
)
STYLE_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('D1D1D1')))
STYLE_LIGHT_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('E1E1E1')))
STYLE_LIGHTER_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('F1F1F1')))

HeaderLine = namedtuple('HeaderLine', ['egy_acronym', 'egy_title'])
OfficialTextLine = namedtuple('OfficialTextLine', ['text'])
LearningUnitYearLine = namedtuple('LearningUnitYearLine', ['luy_acronym', 'luy_title'])
PrerequisiteLine = namedtuple('PrerequisiteLine', ['text', 'operator', 'luy_acronym', 'luy_title'])


class QuerysetToExcel(ABC):
    @abstractmethod
    def get_queryset(self):
        pass

    @abstractmethod
    def to_excel(self):
        pass


class EducationGroupYearLearningUnitsPrerequisitesToExcel(QuerysetToExcel):
    def __init__(self, egy: EducationGroupYear):
        self.egy = egy

    def get_queryset(self):
        return Prerequisite.objects.filter(
            education_group_year=self.egy
        ).prefetch_related(
            Prefetch(
                "prerequisiteitem_set",
                queryset=PrerequisiteItem.objects.order_by(
                    'group_number',
                    'position'
                ).select_related(
                    "learning_unit"
                ).prefetch_related(
                    Prefetch(
                        "learning_unit__learningunityear_set",
                        queryset=LearningUnitYear.objects.filter(academic_year=self.egy.academic_year),
                        to_attr="luys"
                    )
                ),
                to_attr="items"
            )
        ).select_related(
            "learning_unit_year"
        )

    def _to_workbook(self):
        return generate_prerequisites_workbook(self.egy, self.get_queryset())

    def to_excel(self):
        return save_virtual_workbook(self._to_workbook())


def generate_prerequisites_workbook(egy: EducationGroupYear, prerequisites_qs: QuerySet):
    workbook = Workbook(encoding='utf-8')

    excel_lines = _build_excel_lines(egy, prerequisites_qs)
    header, *content = [tuple(line) for line in excel_lines]

    worksheet_data = {
        WORKSHEET_TITLE_KEY: _(XLS_FILENAME),
        HEADER_TITLES_KEY: header,
        CONTENT_KEY: content,
        STYLED_CELLS: _style_cells(prerequisites_qs)
    }

    _build_worksheet(worksheet_data, workbook, 0)

    _merge_cells(prerequisites_qs, workbook)
    _post_style_cell(workbook)

    return workbook


def _build_excel_lines(egy: EducationGroupYear, prerequisite_qs: QuerySet):
    content = []
    content.append(
        HeaderLine(egy_acronym=egy.acronym, egy_title=egy.title)
    )
    content.append(
        OfficialTextLine(text=_("Official"))
    )

    for prerequisite in prerequisite_qs:
        luy = prerequisite.learning_unit_year
        content.append(
            LearningUnitYearLine(luy_acronym=luy.acronym, luy_title=luy.complete_title)
        )

        groups_generator = itertools.groupby(prerequisite.items, key=lambda item: item.group_number)
        for key, group_gen in groups_generator:
            group = list(group_gen)
            for item in group:
                text = (_("has as prerequisite") + " :") if item.group_number == 1 and item.position == 1 else None
                operator = _get_operator(prerequisite, item)
                luy_acronym = _get_item_acronym(item, group)
                content.append(
                    PrerequisiteLine(
                        text=text,
                        operator=operator,
                        luy_acronym=luy_acronym,
                        luy_title=item.learning_unit.luys[0].complete_title
                    )
                )
    return content


def _get_operator(prerequisite: Prerequisite, prerequisite_item: PrerequisiteItem):
    if prerequisite_item.group_number == 1 and prerequisite_item.position == 1:
        return None
    elif prerequisite_item.position == 1:
        return _(prerequisite.main_operator)
    return _(prerequisite.secondary_operator)


def _get_item_acronym(prerequisite_item: PrerequisiteItem, group: list):
    acronym_format = "{acronym}"
    if prerequisite_item.position == 1 and len(group) > 1:
        acronym_format = "({acronym}"
    elif prerequisite_item.position == len(group) and len(group) > 1:
        acronym_format = "{acronym})"
    return acronym_format.format(acronym=prerequisite_item.learning_unit.luys[0].acronym)


def _style_cells(prerequisites_qs: QuerySet):
    luy_acronym_cells = []
    row_index = 3
    for prerequisite in prerequisites_qs:
        luy_acronym_cells.append(
            "A{row_index}".format(row_index=row_index)
        )
        row_index += (len(prerequisite.items) + 1)

    luy_title_cells = []
    row_index = 3
    for prerequisite in prerequisites_qs:
        luy_title_cells.append(
            "B{row_index}".format(row_index=row_index)
        )
        row_index += (len(prerequisite.items) + 1)

    items_cells = []
    row_index = 3
    for prerequisite in prerequisites_qs:
        for i, item in enumerate(prerequisite.items):
            if i % 2 == 0:
                items_cells.append(
                    "C{row_index}".format(row_index=row_index + i + 1)
                )
                items_cells.append(
                    "D{row_index}".format(row_index=row_index + i + 1)
                )
        row_index += len(prerequisite.items) + 1

    return {
        STYLE_NO_GRAY: ["A1", "B1"],
        STYLE_BORDER_BOTTOM: ["A2"],
        STYLE_GRAY: luy_acronym_cells,
        STYLE_LIGHT_GRAY: luy_title_cells,
        STYLE_LIGHTER_GRAY: items_cells
    }


def _merge_cells(prerequisites_qs, workbook):
    worksheet = workbook.worksheets[0]
    row_index = 3
    for prerequisite in prerequisites_qs:
        worksheet.merge_cells(start_row=row_index, end_row=row_index, start_column=2, end_column=4)
        row_index += (len(prerequisite.items) + 1)


def _post_style_cell(workbook: Workbook):
    worksheet = workbook.worksheets[0]

    main_operator = None
    for row_index in range(3, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_index, column=2)

        if not cell.value or cell.value not in (_(AND), _(OR)):
            main_operator = None

        elif main_operator is None:
            main_operator = cell.value

        elif main_operator != cell.value:
            cell.style = Style(
                font=Font(color=RED)
            )
