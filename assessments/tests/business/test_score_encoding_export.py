##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2018 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import datetime

from django.test import TestCase
from openpyxl import Workbook
from openpyxl.styles import Color

from assessments.business.score_encoding_export import _coloring_enrollment_state, FIRST_COL_LEGEND_ENROLLMENT_STATUS, \
    _color_legend, FIRST_ROW_LEGEND_ENROLLMENT_STATUS
from base.tests.factories.academic_year import create_current_academic_year
from base.tests.factories.exam_enrollment import ExamEnrollmentFactory
from base.tests.factories.session_examen import SessionExamFactory
from base.models.enums import number_session, academic_calendar_type
from assessments.business.enrollment_state import ENROLLED_LATE_COLOR, NOT_ENROLLED_COLOR
from base.models.enums import exam_enrollment_state as enrollment_states
from base.tests.factories.academic_calendar import AcademicCalendarFactory
from base.tests.factories.session_exam_calendar import SessionExamCalendarFactory
from base.tests.factories.learning_unit_year import LearningUnitYearFactory

WHITE_RGB = '00000000'
ROW_NUMBER = 1


class XlsTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.academic_year = create_current_academic_year()
        cls.academic_calendar = AcademicCalendarFactory(title="Submission of score encoding - 1",
                                                        academic_year__current=True,
                                                        reference=academic_calendar_type.SCORES_EXAM_SUBMISSION)
        cls.session_exam_calendar = SessionExamCalendarFactory(academic_calendar=cls.academic_calendar,
                                                               number_session=number_session.ONE)
        learning_unit_yr = LearningUnitYearFactory(learning_container_year__academic_year__current=True)
        cls.session_exam = SessionExamFactory(number_session=number_session.ONE, learning_unit_year=learning_unit_yr)

    def setUp(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def test_coloring_not_enrolled(self):
        exam_enrollment = ExamEnrollmentFactory(session_exam=self.session_exam,
                                                enrollment_state=enrollment_states.NOT_ENROLLED)
        _coloring_enrollment_state(self.worksheet,
                                   ROW_NUMBER,
                                   exam_enrollment)

        self.assertEqual(
            self.worksheet.cell(row=ROW_NUMBER, column=FIRST_COL_LEGEND_ENROLLMENT_STATUS).style.fill.fgColor,
            Color(rgb=NOT_ENROLLED_COLOR.lstrip('#')))

    def test_coloring_enrolled_late(self):
        exam_enrollment = ExamEnrollmentFactory(session_exam=self.session_exam,
                                                enrollment_state=enrollment_states.ENROLLED,
                                                date_enrollment=self.academic_calendar.start_date + datetime.timedelta(
                                                    days=1))
        _coloring_enrollment_state(self.worksheet,
                                   ROW_NUMBER,
                                   exam_enrollment)

        self.assertEqual(
            self.worksheet.cell(row=ROW_NUMBER, column=FIRST_COL_LEGEND_ENROLLMENT_STATUS).style.fill.fgColor,
            Color(rgb=ENROLLED_LATE_COLOR.lstrip('#')))

    def test_coloring_normal_enrollment(self):
        exam_enrollment = ExamEnrollmentFactory(session_exam=self.session_exam,
                                                enrollment_state=enrollment_states.ENROLLED,
                                                date_enrollment=self.academic_calendar.start_date)
        _coloring_enrollment_state(self.worksheet, ROW_NUMBER, exam_enrollment)

        self.assertEqual(
            self.worksheet.cell(row=ROW_NUMBER, column=FIRST_COL_LEGEND_ENROLLMENT_STATUS).style.fill.fgColor,
            Color(rgb=WHITE_RGB))

    def test_color_legend(self):
        _color_legend(self.worksheet)
        self.assertEqual(
            self.worksheet.cell(row=FIRST_ROW_LEGEND_ENROLLMENT_STATUS,
                                column=FIRST_COL_LEGEND_ENROLLMENT_STATUS).style.fill.fgColor,
            Color(rgb=ENROLLED_LATE_COLOR.lstrip('#')))
        self.assertEqual(
            self.worksheet.cell(row=FIRST_ROW_LEGEND_ENROLLMENT_STATUS+1,
                                column=FIRST_COL_LEGEND_ENROLLMENT_STATUS).style.fill.fgColor,
            Color(rgb=NOT_ENROLLED_COLOR.lstrip("#")))
